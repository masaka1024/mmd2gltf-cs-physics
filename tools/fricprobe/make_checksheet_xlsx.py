# -*- coding: utf-8 -*-
"""摩擦合成方式の神託チェックシート (Excel)。
   滑る/止まる をドロップダウンで選ぶと、下段の判定表が自動で「一致」を出す。

使い方: python make_checksheet_xlsx.py [出力.xlsx]
"""
import math, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

PAIRS = [("A", 1.0, 0.25), ("B", 0.5, 0.5)]
TANS = [0.35, 0.55, 0.80]
METHODS = ["積 (Bullet2.75)", "幾何平均 (当既定)", "算術平均", "min", "max"]

def mu(f0, f1):
    return [f0 * f1, math.sqrt(f0 * f1), (f0 + f1) / 2.0, min(f0, f1), max(f0, f1)]

CELLS = [(t, f0, f1, tan) for (t, f0, f1) in PAIRS for tan in TANS]

HDR   = PatternFill("solid", fgColor="1F3864")
SUB   = PatternFill("solid", fgColor="D9E2F3")
INPUT = PatternFill("solid", fgColor="FFF2CC")
OKF   = PatternFill("solid", fgColor="C6EFCE")
NGF   = PatternFill("solid", fgColor="F2F2F2")
CAL   = PatternFill("solid", fgColor="EDEDED")
WHITE = Font(color="FFFFFF", bold=True)
B = Border(*[Side(style="thin", color="BFBFBF")] * 4)

def put(ws, r, c, v, fill=None, font=None, align="center", border=True, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border: cell.border = B
    return cell

def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "摩擦合成方式_チェックシート.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "観察シート"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    for col in "CDEF": ws.column_dimensions[col].width = 9
    for col in "GHI":  ws.column_dimensions[col].width = 16
    ws.column_dimensions["J"].width = 34

    r = 1
    put(ws, r, 1, "摩擦合成方式の神託チェックシート", HDR, WHITE, "left"); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.row_dimensions[r].height = 24
    r += 1
    put(ws, r, 1, "傾いた板の上に箱を置いただけのモデル。物理を有効にして 1〜2 秒見て、"
                  "青い箱が赤い出発点マーカーから出ていけば「滑る」、載ったままなら「止まる」を選ぶ。"
                  "黄色いセルがドロップダウンです。  ★MMD は物理中のボーンマーカーを描かないので "
                  "[物理演算]→[常に演算] にして、青い箱そのものを見ること。  "
                  "PMX の場所: " + os.path.abspath(os.path.dirname(__file__)),
        None, None, "left", False, True)
    ws.row_dimensions[r].height = 44
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    hdr_row = r
    for c, t in enumerate(["#", "ファイル", "坂 f0", "箱 f1", "tanθ", "θ",
                           "PMXe (補正OFF)", "PMXe (補正ON)", "MMD 本体", "備考"], start=1):
        put(ws, r, c, t, HDR, WHITE)
    r += 1
    first_data = r
    dv = DataValidation(type="list", formula1='"滑る,止まる"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    for i, (tag, f0, f1, tan) in enumerate(CELLS, start=1):
        put(ws, r, 1, i)
        put(ws, r, 2, "fric_%s_tan%03d.pmx" % (tag, int(round(tan * 100))), None, None, "left")
        put(ws, r, 3, f0); put(ws, r, 4, f1)
        put(ws, r, 5, tan); put(ws, r, 6, "%.1f°" % math.degrees(math.atan(tan)))
        for c in (7, 8, 9):
            cell = put(ws, r, c, None, INPUT)
            dv.add(cell)
        put(ws, r, 10, "装置確認用 (どの方式でも滑るはず)" if (tag == "B" and tan > 0.4) else "", None, None, "left")
        r += 1
    last_data = r - 1

    r += 1
    put(ws, r, 1, "判定表 — 観察した列を選ぶと、一致する方式の行が緑になります", HDR, WHITE, "left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10); r += 1
    put(ws, r, 1, "観察する列:", None, Font(bold=True), "right"); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    sel = put(ws, r, 3, "PMXe (補正OFF)", INPUT, None, "left"); ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    dv2 = DataValidation(type="list", formula1='"PMXe (補正OFF),PMXe (補正ON),MMD 本体"', allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv2); dv2.add(sel)
    sel_ref = "$C$%d" % r
    r += 1
    # INDEX の列番号は範囲内の相対位置 (G:I の 1..3)。絶対列番号ではない。
    col_pick = ("IF(%s=\"PMXe (補正ON)\",2,IF(%s=\"MMD 本体\",3,1))" % (sel_ref, sel_ref))
    # 作業列 K: 選んだ列の観察値を1本に集める。判定式を配列数式にしないための足場。
    put(ws, hdr_row, 11, "観察 (選択列)", SUB, Font(bold=True))
    for ci in range(len(CELLS)):
        rr = first_data + ci
        put(ws, rr, 11, "=IF(INDEX($G%d:$I%d,1,%s)=\"\",\"\",INDEX($G%d:$I%d,1,%s))"
            % (rr, rr, col_pick, rr, rr, col_pick), CAL)
    ws.column_dimensions["K"].width = 13
    r += 1

    jhdr = r
    for c, t in enumerate(["方式", "μ(対A)", "μ(対B)"] + ["#%d %s%.2f" % (i, tag, tan)
                          for i, (tag, f0, f1, tan) in enumerate(CELLS, start=1)] + ["判定"], start=1):
        put(ws, r, c, t, SUB, Font(bold=True))
    r += 1
    first_j = r
    for mi, name in enumerate(METHODS):
        put(ws, r, 1, name, None, Font(bold=(mi < 2)), "left")
        put(ws, r, 2, round(mu(1.0, 0.25)[mi], 4))
        put(ws, r, 3, round(mu(0.5, 0.5)[mi], 4))
        parts = []
        for ci, (tag, f0, f1, tan) in enumerate(CELLS):
            m = mu(f0, f1)[mi]
            pred = "滑る" if tan > m else "止まる"
            put(ws, r, 4 + ci, pred)
            obs = "$K$%d" % (first_data + ci)
            parts.append('IF({o}="",0,IF({o}="{p}",0,1))'.format(o=obs, p=pred))
        # 未記入は無視し、記入ぶんが全部一致なら「一致」。配列数式は使わない。
        rng = "$K$%d:$K$%d" % (first_data, first_data + len(CELLS) - 1)
        filled = 'COUNTIF(%s,"滑る")+COUNTIF(%s,"止まる")' % (rng, rng)
        f = ('=IF(%s=0,"—",IF(%s=0,"★一致","不一致"))'
             % (filled, "+".join(parts)))
        c = put(ws, r, 4 + len(CELLS), f, None, Font(bold=True))
        r += 1
    last_j = r - 1
    jcol = get_column_letter(4 + len(CELLS))
    ws.column_dimensions[jcol].width = 10
    ws.conditional_formatting.add("A%d:%s%d" % (first_j, jcol, last_j),
        FormulaRule(formula=['$%s%d="★一致"' % (jcol, first_j)], fill=OKF))

    r += 1
    put(ws, r, 1, "読み方: #4 (B 0.35) が「滑る」なら 積 で確定。#5 #6 はどの方式でも滑るので装置確認用。"
                  "ここが「止まる」なら物理が動いていないなど観測条件を疑うこと。", None, None, "left", False, True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10); ws.row_dimensions[r].height = 28
    r += 2

    put(ws, r, 1, "参考: 装置の較正 (2026-08-23 実測・120フレーム=2秒の移動量)", HDR, WHITE, "left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10); r += 1
    for c, t in enumerate(["実装"] + ["#%d" % i for i in range(1, len(CELLS) + 1)] + ["期待"], start=1):
        put(ws, r, c, t, SUB, Font(bold=True))
    r += 1
    calib = [("純Bullet (bulletref)", [30.3, 90.6, 125.0, 30.3, 90.6, 125.0], "積"),
             ("当エンジン 既定", [0.02, 8.9, 88.7, 0.02, 8.9, 88.7], "幾何平均"),
             ("当エンジン FRICMUL=1", [31.6, 92.0, 125.2, 31.6, 92.0, 125.2], "積")]
    for name, vals, exp in calib:
        put(ws, r, 1, name, CAL, None, "left")
        for i, v in enumerate(vals):
            put(ws, r, 2 + i, ("%.2f 滑る" % v) if v > 0.5 else ("%.2f 止まる" % v), CAL)
        put(ws, r, 2 + len(vals), exp, CAL)
        r += 1
    r += 1
    put(ws, r, 1, "→ どちらも判定表の予測と完全一致。装置は判別できている。", None, None, "left", False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    ws.freeze_panes = "A%d" % first_data

    # ================= 第2ラウンド =================
    w2 = wb.create_sheet("第2ラウンド")
    w2.column_dimensions["A"].width = 5
    w2.column_dimensions["B"].width = 22
    for col in "CD": w2.column_dimensions[col].width = 9
    w2.column_dimensions["E"].width = 16
    w2.column_dimensions["F"].width = 44
    LAD = [(7, 0.30), (8, 0.22), (9, 0.15), (10, 0.08)]
    r = 1
    put(w2, r, 1, "第2ラウンド — 「全部滑る」の裏取り", HDR, WHITE, "left")
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); w2.row_dimensions[r].height = 24
    r += 1
    put(w2, r, 1, "第1ラウンドは全6件「滑る」だった。候補5方式でこれに当てはまるのは 積 だけ。"
                  "ただし『摩擦がまるで効いていない (μ≈0)』でも同じ結果になるので切り分けが要る。"
                  "浅い角度を並べて、止まり始める角度から μ を直に読む。対B (坂0.5 / 箱0.5) のみ。",
        None, None, "left", False, True)
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); w2.row_dimensions[r].height = 44
    r += 2
    for c, t in enumerate(["#", "ファイル", "tanθ", "θ", "観察", "備考"], start=1):
        put(w2, r, c, t, HDR, WHITE)
    r += 1
    dv3 = DataValidation(type="list", formula1='"滑る,止まる"', allow_blank=True, showDropDown=False)
    w2.add_data_validation(dv3)
    f2 = r
    for (num, tan) in LAD:
        put(w2, r, 1, num)
        put(w2, r, 2, "fric_L_tan%03d.pmx" % int(round(tan * 100)), None, None, "left")
        put(w2, r, 3, tan)
        put(w2, r, 4, "%.1f°" % math.degrees(math.atan(tan)))
        dv3.add(put(w2, r, 5, None, INPUT))
        r += 1
    r += 1
    put(w2, r, 1, "判定", HDR, WHITE, "left")
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    for c, t in enumerate(["観察", "μ", "意味"], start=1):
        put(w2, r, c, t, SUB, Font(bold=True))
    w2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    r += 1
    rows = [("#7 滑る / #8 止まる", "0.22〜0.30", "★積 (0.25) で確定。第1ラウンドの読みが裏付けられる"),
            ("#7#8 滑る / #9 止まる", "0.15〜0.22", "積より小さい。別の規則 (積×何か) を疑う"),
            ("#7#8#9 滑る / #10 止まる", "0.08〜0.15", "かなり小さい。摩擦を弱く扱っている"),
            ("全部 滑る", "< 0.08", "★摩擦がほぼ効いていない。「積」という読みは取り下げ")]
    for a, b_, c_ in rows:
        put(w2, r, 1, a, None, None, "left"); put(w2, r, 2, b_)
        put(w2, r, 3, c_, None, None, "left"); w2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        r += 1
    r += 1
    put(w2, r, 1, "較正 (当エンジンで確認済み・120F の移動量)", HDR, WHITE, "left")
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    for c, t in enumerate(["実装", "#7 0.30", "#8 0.22", "#9 0.15", "#10 0.08"], start=1):
        put(w2, r, c, t, SUB, Font(bold=True))
    r += 1
    for name, vals in (("FRICMUL=1 (積 μ=0.25)", [9.80, 0.16, 0.12, 0.04]),
                       ("既定 (幾何平均 μ=0.5)", [0.11, 0.02, 0.02, 0.05])):
        put(w2, r, 1, name, CAL, None, "left")
        for i, v in enumerate(vals):
            put(w2, r, 2 + i, ("%.2f 滑る" % v) if v > 0.5 else ("%.2f 止まる" % v), CAL)
        r += 1
    r += 1
    put(w2, r, 1, "→ μ=0.25 の境目が 0.22 と 0.30 の間に正しく出ている。装置は μ を挟めている。",
        None, None, "left", False)
    w2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    wb.save(out)
    print("書き出した:", os.path.abspath(out))

main()
